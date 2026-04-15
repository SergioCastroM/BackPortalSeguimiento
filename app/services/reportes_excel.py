"""Generación de archivos .xlsx reales para reportes de administración."""
from __future__ import annotations

import io
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, func
from sqlalchemy.orm import Session, joinedload

from app.models import (
    IndicadorProducto,
    Meta,
    Producto,
    Programa,
    Sector,
    Secretaria,
    SeguimientoMeta,
)
from app.services.dashboard_service import dashboard_global, dashboard_secretaria

_VALOR_ESPERADO_BY_ANIO = {
    2024: "valor_esperado_2024",
    2025: "valor_esperado_2025",
    2026: "valor_esperado_2026",
    2027: "valor_esperado_2027",
}


def _valor_esperado_meta_por_anio(meta: Meta, anio: int) -> float:
    attr = _VALOR_ESPERADO_BY_ANIO.get(anio)
    if not attr:
        return 0.0
    v = getattr(meta, attr, None)
    return float(v or 0)


def _seguimientos_por_trimestre(meta: Meta, anio: int) -> dict[int, SeguimientoMeta]:
    out: dict[int, SeguimientoMeta] = {}
    for s in meta.seguimientos or []:
        if s.anio == anio and 1 <= s.trimestre <= 4:
            out[s.trimestre] = s
    return out


def _cell_seg_num(seg: SeguimientoMeta | None, attr: str) -> float | str:
    if not seg:
        return ""
    v = getattr(seg, attr, None)
    if v is None:
        return ""
    return float(v)


def _cell_seg_text(seg: SeguimientoMeta | None, attr: str) -> str:
    if not seg:
        return ""
    v = getattr(seg, attr, None)
    if v is None:
        return ""
    return str(v).strip()


def _cell_seg_fecha(seg: SeguimientoMeta | None) -> str:
    if not seg or not seg.fecha_registro:
        return ""
    return seg.fecha_registro.isoformat()


def _append_sheet_metas_seguimiento_trimestres(
    wb: Workbook,
    db: Session,
    anio: int,
    bold: Font,
) -> None:
    """Una fila por meta activa; por cada T1–T4 del año: valores ingresados en seguimiento."""
    metas = (
        db.query(Meta)
        .filter(Meta.activo == True)
        .options(
            joinedload(Meta.secretaria),
            joinedload(Meta.indicador_producto)
            .joinedload(IndicadorProducto.producto)
            .joinedload(Producto.programa)
            .joinedload(Programa.sector),
            joinedload(Meta.seguimientos),
        )
        .order_by(Meta.secretaria_id, Meta.id)
        .all()
    )

    ws = wb.create_sheet("Metas y seguimiento")
    base_headers = [
        "ID meta",
        "Secretaría",
        "Descripción",
        "Cód. indicador",
        "Nombre indicador",
        "Sector",
        f"Valor esperado {anio}",
    ]
    per_t_headers = [
        "Valor ejecutado",
        "Recursos ejecutados",
        "% cumplimiento",
        "Evidencia",
        "Observaciones",
        "Fecha registro",
    ]
    header_row: list[str] = list(base_headers)
    for t in (1, 2, 3, 4):
        for h in per_t_headers:
            header_row.append(f"T{t} — {h}")
    ws.append(header_row)
    for c in ws[1]:
        c.font = bold
    ws.freeze_panes = "A2"

    for meta in metas:
        ip = meta.indicador_producto
        sector_n = ""
        if ip and ip.producto and ip.producto.programa and ip.producto.programa.sector:
            sector_n = ip.producto.programa.sector.nombre or ""
        sec_nombre = meta.secretaria.nombre if meta.secretaria else ""
        by_t = _seguimientos_por_trimestre(meta, anio)
        row: list = [
            meta.id,
            sec_nombre,
            meta.descripcion or "",
            (ip.codigo if ip else "") or "",
            (ip.nombre if ip else "") or "",
            sector_n,
            _valor_esperado_meta_por_anio(meta, anio),
        ]
        for t in (1, 2, 3, 4):
            s = by_t.get(t)
            row.extend(
                [
                    _cell_seg_num(s, "valor_ejecutado"),
                    _cell_seg_num(s, "recursos_ejecutados"),
                    _cell_seg_num(s, "porcentaje_cumplimiento"),
                    _cell_seg_text(s, "evidencia"),
                    _cell_seg_text(s, "observaciones"),
                    _cell_seg_fecha(s),
                ]
            )
        ws.append(row)

    # Anchos razonables (muchas columnas; evita autosize pesado)
    widths = [10, 28, 48, 14, 32, 22, 16]
    for t in range(4):
        widths.extend([16, 16, 12, 36, 36, 22])
    for idx, wch in enumerate(widths, start=1):
        if idx <= ws.max_column:
            ws.column_dimensions[get_column_letter(idx)].width = min(wch, 50)


def _autosize_columns(ws, max_width: int = 48) -> None:
    for col in ws.columns:
        letter = col[0].column_letter
        width = min(max(len(str(c.value or "")) for c in col) + 2, max_width)
        ws.column_dimensions[letter].width = width


def build_reporte_total_excel(db: Session, anio: int, trimestre: int) -> bytes:
    data = dashboard_global(db, anio, trimestre)
    wb = Workbook()

    ws0 = wb.active
    ws0.title = "Resumen"
    bold = Font(bold=True)
    ws0["A1"] = f"Consolidado total — Año {anio} · Trimestre {trimestre}"
    ws0["A1"].font = bold
    kpis = data["kpis"]
    rows = [
        ("Total metas activas", kpis["total_metas"]),
        ("Con seguimiento en el trimestre", kpis["con_seguimiento"]),
        ("Pendientes", kpis["pendientes"]),
        ("% cumplimiento promedio", kpis["porcentaje_cumplimiento_prom"]),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        ws0.cell(row=i, column=1, value=k).font = bold
        ws0.cell(row=i, column=2, value=v)
    _autosize_columns(ws0)

    ws1 = wb.create_sheet("Por secretaría")
    ws1.append(["ID", "Secretaría", "Total metas", "% cumplimiento (prom.)"])
    for c in ws1[1]:
        c.font = bold
    for row in data["por_secretaria"]:
        ws1.append(
            [
                row["secretaria_id"],
                row["secretaria_nombre"],
                row["total_metas"],
                round(row["porcentaje"], 2),
            ]
        )
    _autosize_columns(ws1)

    ws2 = wb.create_sheet("Por sector")
    ws2.append(["ID", "Sector", "Cantidad metas", "% del total"])
    for c in ws2[1]:
        c.font = bold
    for row in data["por_sector"]:
        ws2.append(
            [
                row["sector_id"],
                row["sector_nombre"],
                row["cantidad"],
                row["porcentaje"],
            ]
        )
    _autosize_columns(ws2)

    # % de metas activas con seguimiento en cada trimestre (filas = secretarías)
    ws3 = wb.create_sheet("Cobertura por trimestre")
    by_sec: dict[int, dict] = {}
    for h in data["heatmap"]:
        sid = h["secretaria_id"]
        if sid not in by_sec:
            by_sec[sid] = {"nombre": h["secretaria_nombre"], "t": {}}
        by_sec[sid]["t"][h["trimestre"]] = h["porcentaje"]
    ws3.append(["Secretaría", "T1", "T2", "T3", "T4"])
    for c in ws3[1]:
        c.font = bold
    for sid in sorted(by_sec.keys()):
        info = by_sec[sid]
        ws3.append(
            [
                info["nombre"],
                info["t"].get(1) if info["t"].get(1) is not None else "",
                info["t"].get(2) if info["t"].get(2) is not None else "",
                info["t"].get(3) if info["t"].get(3) is not None else "",
                info["t"].get(4) if info["t"].get(4) is not None else "",
            ]
        )
    _autosize_columns(ws3)

    _append_sheet_metas_seguimiento_trimestres(wb, db, anio, bold)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_reporte_secretaria_excel(
    db: Session, secretaria_id: int, anio: int, trimestre: int
) -> bytes:
    data = dashboard_secretaria(db, secretaria_id, anio, trimestre)
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    bold = Font(bold=True)
    sec = data["secretaria"]
    ws["A1"] = f"Reporte secretaría: {sec.get('nombre', '')} — {anio} T{trimestre}"
    ws["A1"].font = bold
    k = data["kpis"]
    for i, (label, val) in enumerate(
        [
            ("Total metas", k["total_metas"]),
            ("Registradas en el trimestre", k["registradas"]),
            ("Pendientes", k["pendientes"]),
            ("% cumplimiento", k["porcentaje_cumplimiento"]),
        ],
        start=3,
    ):
        ws.cell(row=i, column=1, value=label).font = bold
        ws.cell(row=i, column=2, value=val)
    _autosize_columns(ws)

    ws2 = wb.create_sheet("Metas")
    ws2.append(
        [
            "ID meta",
            "Descripción",
            "Indicador",
            "Sector",
            "Valor esperado 2026",
            f"% cumpl. T{trimestre}",
            "Valor ejecutado",
        ]
    )
    for c in ws2[1]:
        c.font = bold
    for m in data["metas"]:
        ip = m.get("indicador_producto") or {}
        sector = ""
        if ip.get("producto") and ip["producto"].get("programa"):
            pr = ip["producto"]["programa"]
            if pr.get("sector"):
                sector = pr["sector"].get("nombre") or ""
        seg_t = next(
            (s for s in m.get("seguimientos", []) if s["anio"] == anio and s["trimestre"] == trimestre),
            None,
        )
        ws2.append(
            [
                m["id"],
                (m.get("descripcion") or "")[:500],
                ip.get("nombre") or ip.get("codigo") or "",
                sector,
                m.get("valor_esperado_2026", 0),
                float(seg_t["porcentaje_cumplimiento"]) if seg_t else "",
                "",
            ]
        )
    # Valor ejecutado: requiere datos de ORM; usar metas_esperado_vs_ejecutado
    ev_by_id = {x["meta_id"]: x for x in data.get("metas_esperado_vs_ejecutado", [])}
    for row_idx in range(2, ws2.max_row + 1):
        mid = ws2.cell(row=row_idx, column=1).value
        if mid in ev_by_id:
            ws2.cell(row=row_idx, column=7, value=ev_by_id[mid]["ejecutado"])
    _autosize_columns(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_reporte_sector_excel(db: Session, sector_id: int, anio: int, trimestre: int) -> bytes:
    sector = db.get(Sector, sector_id)
    nombre_sec = sector.nombre if sector else f"ID {sector_id}"

    q = (
        db.query(Meta, Secretaria.nombre, SeguimientoMeta)
        .join(Secretaria, Meta.secretaria_id == Secretaria.id)
        .join(IndicadorProducto, Meta.indicador_producto_id == IndicadorProducto.id)
        .join(Producto, IndicadorProducto.producto_id == Producto.id)
        .join(Programa, Producto.programa_id == Programa.id)
        .outerjoin(
            SeguimientoMeta,
            and_(
                SeguimientoMeta.meta_id == Meta.id,
                SeguimientoMeta.anio == anio,
                SeguimientoMeta.trimestre == trimestre,
            ),
        )
        .filter(Meta.activo == True, Programa.sector_id == sector_id)
        .order_by(Secretaria.nombre, Meta.id)
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Metas"
    bold = Font(bold=True)
    ws["A1"] = f"Sector: {nombre_sec} — {anio} T{trimestre}"
    ws["A1"].font = bold
    ws.append([])
    ws.append(
        [
            "ID meta",
            "Secretaría",
            "Descripción",
            "Indicador",
            "% cumplimiento",
            "Valor ejecutado",
        ]
    )
    for c in ws[3]:
        c.font = bold
    for meta, sec_nombre, seg in q.all():
        ip = meta.indicador_producto
        ind_label = ""
        if ip:
            ind_label = ip.nombre or ip.codigo or ""
        ws.append(
            [
                meta.id,
                sec_nombre,
                (meta.descripcion or "")[:500],
                ind_label,
                float(seg.porcentaje_cumplimiento or 0) if seg else "",
                float(seg.valor_ejecutado or 0) if seg else "",
            ]
        )
    _autosize_columns(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_reporte_pendientes_excel(db: Session, anio: int, trimestre: int) -> bytes:
    con_seg_subq = (
        db.query(SeguimientoMeta.meta_id)
        .filter(
            SeguimientoMeta.anio == anio,
            SeguimientoMeta.trimestre == trimestre,
        )
        .distinct()
    )
    metas = (
        db.query(Meta)
        .filter(Meta.activo == True, ~Meta.id.in_(con_seg_subq))
        .options(
            joinedload(Meta.secretaria),
            joinedload(Meta.indicador_producto)
            .joinedload(IndicadorProducto.producto)
            .joinedload(Producto.programa)
            .joinedload(Programa.sector),
        )
        .order_by(Meta.secretaria_id, Meta.id)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = f"Pendientes T{trimestre}"
    bold = Font(bold=True)
    ws["A1"] = f"Metas sin seguimiento — {anio} · Trimestre {trimestre}"
    ws["A1"].font = bold
    ws.append([])
    ws.append(["ID", "Secretaría", "Descripción", "Indicador", "Sector"])
    for c in ws[3]:
        c.font = bold
    for m in metas:
        sec_n = m.secretaria.nombre if m.secretaria else ""
        ip = m.indicador_producto
        ind = (ip.nombre or ip.codigo or "") if ip else ""
        sector_n = ""
        if ip and ip.producto and ip.producto.programa and ip.producto.programa.sector:
            sector_n = ip.producto.programa.sector.nombre or ""
        ws.append([m.id, sec_n, (m.descripcion or "")[:500], ind, sector_n])
    _autosize_columns(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
